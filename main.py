import customtkinter as ctk
from datetime import datetime, timedelta
import database as db
import sqlite3
from tkinter import filedialog, messagebox
import os

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

DEFAULT_CALORIE_TARGET = 2000
DEFAULT_WATER_TARGET = 2.5

class FitnessDashboard(ctk.CTk):
    def __init__(self):
        super().__init__()
        db.init_db()
        self.ensure_extra_columns_and_tables()
        
        self.title("30-Day Fitness & Calorie Command Center")
        self.geometry("1350x920")
        
        # Temiz kapanış protokolü (asılı kalma hatasını çözer)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # İkon dosyası varsa pencereye ekle
        if os.path.exists("app_icon.ico"):
            try:
                self.iconbitmap("app_icon.ico")
            except Exception: pass
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        # --- Sidebar ---
        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        self.logo = ctk.CTkLabel(self.sidebar, text="🔥 CORE DASHBOARD", font=ctk.CTkFont(size=18, weight="bold"))
        self.logo.pack(pady=20, padx=10)
        
        self.btn_dash = ctk.CTkButton(self.sidebar, text="🏠 Ana Dashboard", command=self.show_dashboard, fg_color="transparent", anchor="w")
        self.btn_dash.pack(fill="x", pady=5, padx=10)
        
        self.btn_workout = ctk.CTkButton(self.sidebar, text="🏃 Egzersiz & Yürüyüş", command=self.show_combined_activity, fg_color="transparent", anchor="w")
        self.btn_workout.pack(fill="x", pady=5, padx=10)

        self.btn_weight = ctk.CTkButton(self.sidebar, text="⚖️ Kilo & VKI Takibi", command=self.show_weight_tracker, fg_color="transparent", anchor="w")
        self.btn_weight.pack(fill="x", pady=5, padx=10)

        self.btn_manage_ex = ctk.CTkButton(self.sidebar, text="⚙️ Egzersiz Kütüphanesi", command=self.show_exercise_manager, fg_color="transparent", anchor="w")
        self.btn_manage_ex.pack(fill="x", pady=5, padx=10)

        self.btn_analytics = ctk.CTkButton(self.sidebar, text="📊 Analiz & Raporlar", command=self.show_analytics, fg_color="transparent", anchor="w")
        self.btn_analytics.pack(fill="x", pady=5, padx=10)

        self.btn_settings = ctk.CTkButton(self.sidebar, text="🎯 Hedef Ayarları", command=self.show_settings, fg_color="transparent", anchor="w")
        self.btn_settings.pack(fill="x", pady=5, padx=10)

        # Excel Butonları
        self.btn_export = ctk.CTkButton(self.sidebar, text="📊 Excel'e Aktar (.xlsx)", command=self.export_to_excel, fg_color="#27ae60", hover_color="#219150", anchor="w")
        self.btn_export.pack(fill="x", pady=(20, 5), padx=10)

        self.btn_import = ctk.CTkButton(self.sidebar, text="📥 Excel'den Yükle", command=self.import_from_excel, fg_color="#2980b9", hover_color="#1c5980", anchor="w")
        self.btn_import.pack(fill="x", pady=5, padx=10)

        # Main Frame
        self.main_frame = ctk.CTkScrollableFrame(self, corner_radius=15, fg_color="#1a1a1a")
        self.main_frame.grid(row=0, column=1, sticky="nsew", padx=15, pady=15)
        
        self.show_dashboard()

    def on_closing(self):
        """Pencere kapatıldığında arka plan zamanlayıcılarını temiz şekilde sonlandırır."""
        try:
            self.quit()
            self.destroy()
        except Exception:
            pass

    def ensure_extra_columns_and_tables(self):
        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        try: cursor.execute("ALTER TABLE daily_logs ADD COLUMN weight REAL")
        except sqlite3.OperationalError: pass
        try: cursor.execute("ALTER TABLE daily_logs ADD COLUMN note TEXT")
        except sqlite3.OperationalError: pass

        cursor.execute('''CREATE TABLE IF NOT EXISTS user_settings (
                            id INTEGER PRIMARY KEY DEFAULT 1,
                            target_calories INTEGER,
                            target_water REAL)''')
        
        cursor.execute("SELECT target_calories, target_water FROM user_settings WHERE id = 1")
        if not cursor.fetchone():
            cursor.execute("INSERT INTO user_settings (id, target_calories, target_water) VALUES (1, ?, ?)",
                           (DEFAULT_CALORIE_TARGET, DEFAULT_WATER_TARGET))
        conn.commit()
        conn.close()

    def get_user_targets(self):
        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT target_calories, target_water FROM user_settings WHERE id = 1")
        row = cursor.fetchone()
        conn.close()
        return (row[0] if row and row[0] else DEFAULT_CALORIE_TARGET, 
                row[1] if row and row[1] else DEFAULT_WATER_TARGET)

    def clear_main_frame(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()

    def get_today_data(self):
        today = datetime.now().strftime("%Y-%m-%d")
        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT walk_calories, water_l, weight, note FROM daily_logs WHERE date = ?", (today,))
        row = cursor.fetchone()
        cursor.execute("SELECT SUM(total_calories) FROM workout_logs WHERE date = ? AND completed = 1", (today,))
        ex_row = cursor.fetchone()
        conn.close()

        walk_cal = row[0] if row and row[0] else 0
        water = row[1] if row and row[1] else 0.0
        weight = row[2] if row and row[2] else 0.0
        note = row[3] if row and row[3] else ""
        ex_cal = int(ex_row[0]) if ex_row and ex_row[0] else 0
        return walk_cal + ex_cal, walk_cal, ex_cal, water, weight, note

    # --- 1. DASHBOARD ---
    def show_dashboard(self):
        self.clear_main_frame()
        today = datetime.now().strftime("%Y-%m-%d")
        total_act, walk_cal, ex_cal, water, weight, note = self.get_today_data()
        target_cal, target_water = self.get_user_targets()
        
        ctk.CTkLabel(self.main_frame, text="Günün Özeti & İlerleme", font=ctk.CTkFont(size=24, weight="bold")).pack(anchor="w", pady=10, padx=10)

        grid_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        grid_frame.pack(fill="x", pady=5)
        grid_frame.columnconfigure((0,1,2,3), weight=1)

        cal_ratio = min(total_act / float(target_cal), 1.0)
        self.create_progress_metric_card(grid_frame, 0, 0, "🔥 Toplam Kalori", f"{total_act} / {target_cal} kcal", cal_ratio, "#e74c3c")
        self.create_metric_card(grid_frame, 0, 1, "🚶 Yürüyüş Kalorisi", f"{walk_cal} kcal", "#e67e22")
        self.create_metric_card(grid_frame, 0, 2, "💪 Antrenman", f"{ex_cal} kcal", "#9b59b6")
        
        water_ratio = min(water / float(target_water), 1.0)
        self.create_progress_metric_card(grid_frame, 0, 3, "💧 Su İlerlemesi", f"{round(water, 2)} / {target_water} L", water_ratio, "#3498db")

        water_bar = ctk.CTkFrame(self.main_frame, fg_color="#2b2b2b", corner_radius=10)
        water_bar.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(water_bar, text="💧 Hızlı Su Kaydı:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10, pady=10)

        def add_water(amount):
            db.log_daily_data(today, {"water_l": round(water + amount, 2)})
            self.show_dashboard()

        ctk.CTkButton(water_bar, text="+250 ml", width=80, fg_color="#3498db", command=lambda: add_water(0.25)).pack(side="left", padx=5)
        ctk.CTkButton(water_bar, text="+500 ml", width=80, fg_color="#2980b9", command=lambda: add_water(0.50)).pack(side="left", padx=5)
        ctk.CTkButton(water_bar, text="Sıfırla", width=60, fg_color="#7f8c8d", command=lambda: (db.log_daily_data(today, {"water_l": 0.0}), self.show_dashboard())).pack(side="left", padx=5)

        note_frame = ctk.CTkFrame(self.main_frame, fg_color="#2b2b2b", corner_radius=10)
        note_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(note_frame, text="📝 Günlük Not / Antrenman Modu:", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(8, 2))
        
        e_note = ctk.CTkEntry(note_frame, placeholder_text="Örn: Enerjik hissettim...", width=600)
        e_note.insert(0, note)
        e_note.pack(side="left", padx=10, pady=(0, 10), fill="x", expand=True)

        def save_note():
            db.log_daily_data(today, {"note": e_note.get()})
            messagebox.showinfo("Başarılı", "Günlük not kaydedildi!")

        ctk.CTkButton(note_frame, text="Kaydet", width=80, fg_color="#8e44ad", command=save_note).pack(side="right", padx=10, pady=(0, 10))

        # Lazy Import ile Grafik Çizimi
        ctk.CTkLabel(self.main_frame, text="📈 Son 7 Günlük Kalori Yakımı", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w", pady=(15, 5), padx=10)
        today_date = datetime.now()
        dates, total_cals = [], []
        
        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        for i in range(6, -1, -1):
            d_str = (today_date - timedelta(days=i)).strftime("%Y-%m-%d")
            dates.append(d_str[5:])
            cursor.execute("SELECT walk_calories FROM daily_logs WHERE date = ?", (d_str,))
            w_row = cursor.fetchone()
            w_c = w_row[0] if w_row and w_row[0] else 0

            cursor.execute("SELECT SUM(total_calories) FROM workout_logs WHERE date = ? AND completed = 1", (d_str,))
            e_row = cursor.fetchone()
            e_c = e_row[0] if e_row and e_row[0] else 0
            total_cals.append(w_c + e_c)
        conn.close()

        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        
        fig, ax = plt.subplots(figsize=(8, 3.0), facecolor="#1a1a1a")
        ax.set_facecolor("#1a1a1a")
        bars = ax.bar(dates, total_cals, color="#e74c3c", width=0.45)
        ax.set_ylabel("Kalori (kcal)", color="white", fontsize=9)
        ax.tick_params(colors='white', labelsize=8)
        for spine in ax.spines.values(): spine.set_color('#444444')

        for bar in bars:
            yval = bar.get_height()
            if yval > 0:
                ax.text(bar.get_x() + bar.get_width()/2.0, yval + 5, f"{int(yval)}", ha='center', va='bottom', color='white', fontsize=8)

        canvas = FigureCanvasTkAgg(fig, master=self.main_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="x", padx=10, pady=10)

    def create_metric_card(self, parent, row, col, title, value, color):
        card = ctk.CTkFrame(parent, fg_color="#2b2b2b", corner_radius=10)
        card.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
        ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=12), text_color="#aaaaaa").pack(anchor="w", padx=10, pady=(10, 2))
        ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=18, weight="bold"), text_color=color).pack(anchor="w", padx=10, pady=(0, 10))

    def create_progress_metric_card(self, parent, row, col, title, value, ratio, color):
        card = ctk.CTkFrame(parent, fg_color="#2b2b2b", corner_radius=10)
        card.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
        ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=12), text_color="#aaaaaa").pack(anchor="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=16, weight="bold"), text_color=color).pack(anchor="w", padx=10, pady=0)
        
        p_bar = ctk.CTkProgressBar(card, progress_color=color, width=150, height=8)
        p_bar.set(ratio)
        p_bar.pack(anchor="w", padx=10, pady=(6, 2))
        ctk.CTkLabel(card, text=f"%{int(ratio*100)} Tamamlandı", font=ctk.CTkFont(size=10), text_color="#888888").pack(anchor="w", padx=10, pady=(0, 6))

    # --- HEDEF AYARLARI ---
    def show_settings(self):
        self.clear_main_frame()
        ctk.CTkLabel(self.main_frame, text="🎯 Günlük Hedef Ayarları", font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=10, padx=10)
        t_cal, t_water = self.get_user_targets()

        card = ctk.CTkFrame(self.main_frame, fg_color="#2b2b2b", corner_radius=10)
        card.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(card, text="Günlük Kalori Hedefi (kcal):", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=15, pady=(15, 2))
        e_cal = ctk.CTkEntry(card, placeholder_text=f"Varsayılan: {DEFAULT_CALORIE_TARGET}", width=200)
        e_cal.insert(0, str(t_cal))
        e_cal.pack(anchor="w", padx=15, pady=(0, 10))

        ctk.CTkLabel(card, text="Günlük Su Hedefi (Litre):", font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=15, pady=(5, 2))
        e_water = ctk.CTkEntry(card, placeholder_text=f"Varsayılan: {DEFAULT_WATER_TARGET}", width=200)
        e_water.insert(0, str(t_water))
        e_water.pack(anchor="w", padx=15, pady=(0, 15))

        def save_settings():
            try:
                c_val = int(e_cal.get()) if e_cal.get() else DEFAULT_CALORIE_TARGET
                w_val = float(e_water.get()) if e_water.get() else DEFAULT_WATER_TARGET
                
                conn = sqlite3.connect(db.DB_NAME)
                cursor = conn.cursor()
                cursor.execute("UPDATE user_settings SET target_calories = ?, target_water = ? WHERE id = 1", (c_val, w_val))
                conn.commit()
                conn.close()

                messagebox.showinfo("Başarılı", "Hedefleriniz başarıyla güncellendi!")
                self.show_dashboard()
            except ValueError:
                messagebox.showerror("Hata", "Lütfen geçerli sayısal değerler girin.")

        def reset_defaults():
            e_cal.delete(0, 'end')
            e_cal.insert(0, str(DEFAULT_CALORIE_TARGET))
            e_water.delete(0, 'end')
            e_water.insert(0, str(DEFAULT_WATER_TARGET))
            save_settings()

        btn_box = ctk.CTkFrame(card, fg_color="transparent")
        btn_box.pack(anchor="w", padx=15, pady=(0, 15))
        ctk.CTkButton(btn_box, text="Hedefleri Kaydet", fg_color="#27ae60", command=save_settings).pack(side="left", padx=(0, 10))
        ctk.CTkButton(btn_box, text="Varsayılana Dön (2000 kcal / 2.5 L)", fg_color="#7f8c8d", command=reset_defaults).pack(side="left")

    # --- EGZERSİZ VE YÜRÜYÜŞ ---
    def show_combined_activity(self):
        self.clear_main_frame()
        today = datetime.now().strftime("%Y-%m-%d")

        ctk.CTkLabel(self.main_frame, text="🚶 Günlük Yürüyüş Kaydı", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w", pady=(10, 5), padx=10)
        walk_frame = ctk.CTkFrame(self.main_frame, fg_color="#2b2b2b", corner_radius=10)
        walk_frame.pack(fill="x", padx=10, pady=5)

        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT walk_minutes, walk_steps, walk_calories FROM daily_logs WHERE date = ?", (today,))
        w_row = cursor.fetchone()
        conn.close()

        e_walk_min = ctk.CTkEntry(walk_frame, placeholder_text="Süre (dk)", width=120)
        e_walk_min.insert(0, str(w_row[0]) if w_row and w_row[0] else "")
        e_walk_min.pack(side="left", padx=10, pady=10)

        e_walk_steps = ctk.CTkEntry(walk_frame, placeholder_text="Adım Sayısı", width=120)
        e_walk_steps.insert(0, str(w_row[1]) if w_row and w_row[1] else "")
        e_walk_steps.pack(side="left", padx=10, pady=10)

        e_walk_cal = ctk.CTkEntry(walk_frame, placeholder_text="Kalori (kcal)", width=120)
        e_walk_cal.insert(0, str(w_row[2]) if w_row and w_row[2] else "")
        e_walk_cal.pack(side="left", padx=10, pady=10)

        def save_walk():
            data = {}
            if e_walk_min.get(): data["walk_minutes"] = int(e_walk_min.get())
            if e_walk_steps.get(): data["walk_steps"] = int(e_walk_steps.get())
            if e_walk_cal.get(): data["walk_calories"] = int(e_walk_cal.get())
            if data: db.log_daily_data(today, data)

        ctk.CTkButton(walk_frame, text="Yürüyüşü Kaydet", fg_color="#2980b9", width=130, command=save_walk).pack(side="right", padx=10, pady=10)

        ctk.CTkLabel(self.main_frame, text="🏃 Bugünkü Egzersiz Programın", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w", pady=(20, 5), padx=10)

        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT exercise_name, sets, reps, cal_per_rep, completed FROM workout_logs WHERE date = ?", (today,))
        today_exs = cursor.fetchall()

        if not today_exs:
            cursor.execute("SELECT name, sets, reps, cal_per_rep FROM exercises")
            for name, s, r, c in cursor.fetchall():
                cursor.execute("INSERT OR IGNORE INTO workout_logs (date, exercise_name, sets, reps, cal_per_rep, total_calories, completed) VALUES (?, ?, ?, ?, ?, ?, 0)",
                               (today, name, s, r, c, int(s*r*c)))
            conn.commit()
            cursor.execute("SELECT exercise_name, sets, reps, cal_per_rep, completed FROM workout_logs WHERE date = ?", (today,))
            today_exs = cursor.fetchall()

        for name, sets_val, reps_val, cal_rep_val, completed_val in today_exs:
            row_color = "#1e4620" if completed_val == 1 else ("#4a1515" if completed_val == -1 else "#2b2b2b")
            row = ctk.CTkFrame(self.main_frame, fg_color=row_color, corner_radius=8)
            row.pack(fill="x", padx=10, pady=5)

            ctk.CTkLabel(row, text=name, font=ctk.CTkFont(size=14, weight="bold"), width=150, anchor="w").pack(side="left", padx=10, pady=10)

            e_sets = ctk.CTkEntry(row, width=45)
            e_sets.insert(0, str(sets_val))
            e_sets.pack(side="left", padx=2)
            ctk.CTkLabel(row, text="Set", text_color="#aaaaaa").pack(side="left", padx=(0, 5))

            e_reps = ctk.CTkEntry(row, width=45)
            e_reps.insert(0, str(reps_val))
            e_reps.pack(side="left", padx=2)
            ctk.CTkLabel(row, text="Tekrar", text_color="#aaaaaa").pack(side="left", padx=(0, 5))

            e_cal_rep = ctk.CTkEntry(row, width=50)
            e_cal_rep.insert(0, str(cal_rep_val))
            e_cal_rep.pack(side="left", padx=2)
            ctk.CTkLabel(row, text="kcal/tkn", text_color="#aaaaaa").pack(side="left", padx=(0, 8))

            init_tot = int(float(sets_val) * float(reps_val) * float(cal_rep_val))
            lbl_tot_cal = ctk.CTkLabel(row, text=f"🔥 {init_tot} kcal", font=ctk.CTkFont(weight="bold"), text_color="#e74c3c", width=90)
            lbl_tot_cal.pack(side="left", padx=5)

            def make_mark_func(ex_name, status, frame, es, er, ec, lbl):
                def mark_ex():
                    try:
                        s, r, c = float(es.get()), float(er.get()), float(ec.get())
                        calc = int(s * r * c)
                    except ValueError: s, r, c, calc = 0, 0, 0, 0
                    
                    c_conn = sqlite3.connect(db.DB_NAME)
                    c_cursor = c_conn.cursor()
                    c_cursor.execute('''INSERT OR REPLACE INTO workout_logs (date, exercise_name, sets, reps, cal_per_rep, total_calories, completed)
                                        VALUES (?, ?, ?, ?, ?, ?, ?)''', (today, ex_name, s, r, c, calc, status))
                    c_conn.commit()
                    c_conn.close()
                    frame.configure(fg_color="#1e4620" if status == 1 else "#4a1515")
                    lbl.configure(text=f"🔥 {calc} kcal")
                return mark_ex

            ctk.CTkButton(row, text="❌", width=40, fg_color="#c0392b", command=make_mark_func(name, -1, row, e_sets, e_reps, e_cal_rep, lbl_tot_cal)).pack(side="right", padx=3)
            ctk.CTkButton(row, text="✅", width=40, fg_color="#27ae60", command=make_mark_func(name, 1, row, e_sets, e_reps, e_cal_rep, lbl_tot_cal)).pack(side="right", padx=3)

        conn.close()

    # --- KİLO VE VKI TAKİBİ ---
    def show_weight_tracker(self):
        self.clear_main_frame()
        today = datetime.now().strftime("%Y-%m-%d")

        ctk.CTkLabel(self.main_frame, text="⚖️ Günlük Kilo & VKI (BMI) Takibi", font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=10, padx=10)

        input_frame = ctk.CTkFrame(self.main_frame, fg_color="#2b2b2b", corner_radius=10)
        input_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="Boy (cm):", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10, pady=10)
        e_height = ctk.CTkEntry(input_frame, placeholder_text="Örn: 175", width=80)
        e_height.insert(0, "175")
        e_height.pack(side="left", padx=5)

        ctk.CTkLabel(input_frame, text="Bugünkü Kilo (kg):", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10, pady=10)
        e_weight = ctk.CTkEntry(input_frame, placeholder_text="Örn: 75.5", width=80)
        
        _, _, _, _, current_w, _ = self.get_today_data()
        if current_w > 0: e_weight.insert(0, str(current_w))
        e_weight.pack(side="left", padx=5)

        lbl_bmi_result = ctk.CTkLabel(input_frame, text="VKI: --", font=ctk.CTkFont(size=14, weight="bold"), text_color="#3498db")
        lbl_bmi_result.pack(side="left", padx=20)

        def save_and_calc_bmi():
            try:
                h_m = float(e_height.get()) / 100.0
                w_kg = float(e_weight.get())
                bmi = round(w_kg / (h_m * h_m), 1)
                
                db.log_daily_data(today, {"weight": w_kg})
                lbl_bmi_result.configure(text=f"VKI: {bmi} ({self.get_bmi_status(bmi)})")
                self.show_weight_tracker()
            except ValueError:
                messagebox.showerror("Hata", "Lütfen geçerli boy ve kilo değerleri girin.")

        ctk.CTkButton(input_frame, text="Kaydet & Hesapla", fg_color="#27ae60", command=save_and_calc_bmi).pack(side="left", padx=10)

        ctk.CTkLabel(self.main_frame, text="📉 Kilo Değişim Trendi", font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", pady=(15, 5), padx=10)

        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT date, weight FROM daily_logs WHERE weight IS NOT NULL AND weight > 0 ORDER BY date ASC LIMIT 15")
        rows = cursor.fetchall()
        conn.close()

        if rows:
            dates = [r[0][5:] for r in rows]
            weights = [r[1] for r in rows]

            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            fig, ax = plt.subplots(figsize=(8, 3.2), facecolor="#1a1a1a")
            ax.set_facecolor("#1a1a1a")
            ax.plot(dates, weights, marker='o', color='#3498db', linewidth=2, markersize=6)
            ax.set_ylabel("Kilo (kg)", color="white", fontsize=9)
            ax.tick_params(colors='white', labelsize=8)
            for spine in ax.spines.values(): spine.set_color('#444444')

            canvas = FigureCanvasTkAgg(fig, master=self.main_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="x", padx=10, pady=10)

    def get_bmi_status(self, bmi):
        if bmi < 18.5: return "Zayıf"
        elif 18.5 <= bmi < 24.9: return "Normal"
        elif 25 <= bmi < 29.9: return "Fazla Kilolu"
        else: return "Obez"

    # --- EGZERSİZ KÜTÜPHANESİ ---
    def show_exercise_manager(self):
        self.clear_main_frame()
        ctk.CTkLabel(self.main_frame, text="⚙️ Egzersiz Kütüphanesi Yönetimi", font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=10, padx=10)

        form_frame = ctk.CTkFrame(self.main_frame, fg_color="#2b2b2b", corner_radius=10)
        form_frame.pack(fill="x", padx=10, pady=10)
        
        lbl_form_title = ctk.CTkLabel(form_frame, text="Yeni Egzersiz Tanımla", font=ctk.CTkFont(weight="bold"))
        lbl_form_title.pack(anchor="w", padx=10, pady=5)
        
        f_inner = ctk.CTkFrame(form_frame, fg_color="transparent")
        f_inner.pack(fill="x", padx=10, pady=5)

        e_name = ctk.CTkEntry(f_inner, placeholder_text="Egzersiz Adı", width=150)
        e_name.pack(side="left", padx=5)
        e_sets = ctk.CTkEntry(f_inner, placeholder_text="Set", width=60)
        e_sets.pack(side="left", padx=5)
        e_reps = ctk.CTkEntry(f_inner, placeholder_text="Tekrar", width=60)
        e_reps.pack(side="left", padx=5)
        e_cal = ctk.CTkEntry(f_inner, placeholder_text="kcal/tekrar", width=90)
        e_cal.pack(side="left", padx=5)

        editing_id = {"id": None}  # Düzenleme modunu takip etmek için mutable state

        def save_or_update_ex():
            name_val = e_name.get().strip()
            if not name_val or not e_sets.get() or not e_reps.get() or not e_cal.get():
                messagebox.showerror("Hata", "Lütfen tüm alanları doldurun.")
                return

            try:
                s_val = int(e_sets.get())
                r_val = int(e_reps.get())
                c_val = float(e_cal.get())
            except ValueError:
                messagebox.showerror("Hata", "Set, tekrar ve kalori için geçerli sayılar girin.")
                return

            conn = sqlite3.connect(db.DB_NAME)
            cursor = conn.cursor()

            if editing_id["id"] is None:
                # Yeni Kayıt
                try:
                    cursor.execute("INSERT INTO exercises (name, sets, reps, cal_per_rep) VALUES (?, ?, ?, ?)",
                                   (name_val, s_val, r_val, c_val))
                    conn.commit()
                except sqlite3.IntegrityError:
                    messagebox.showerror("Hata", "Bu isimde bir egzersiz zaten mevcut.")
            else:
                # Güncelleme
                cursor.execute("UPDATE exercises SET name = ?, sets = ?, reps = ?, cal_per_rep = ? WHERE id = ?",
                               (name_val, s_val, r_val, c_val, editing_id["id"]))
                conn.commit()

            conn.close()
            self.show_exercise_manager()

        btn_save = ctk.CTkButton(f_inner, text="Kütüphaneye Kaydet", fg_color="#27ae60", command=save_or_update_ex)
        btn_save.pack(side="left", padx=10)

        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, sets, reps, cal_per_rep FROM exercises")
        all_exs = cursor.fetchall()
        conn.close()

        for ex_id, name, s, r, c in all_exs:
            row = ctk.CTkFrame(self.main_frame, fg_color="#222222", corner_radius=8)
            row.pack(fill="x", padx=10, pady=4)
            ctk.CTkLabel(row, text=name, font=ctk.CTkFont(weight="bold"), width=180, anchor="w").pack(side="left", padx=10, pady=8)
            ctk.CTkLabel(row, text=f"Varsayılan: {s} Set x {r} Tekrar | Birim: {c} kcal/tekrar", text_color="#aaaaaa").pack(side="left", padx=10)

            def delete_ex(id_to_del):
                c_conn = sqlite3.connect(db.DB_NAME)
                c_cursor = c_conn.cursor()
                c_cursor.execute("DELETE FROM exercises WHERE id = ?", (id_to_del,))
                c_conn.commit()
                c_conn.close()
                self.show_exercise_manager()

            def edit_ex(id_to_edit, curr_name, curr_s, curr_r, curr_c):
                editing_id["id"] = id_to_edit
                lbl_form_title.configure(text=f"✏️ Egzersiz Düzenle: {curr_name}")
                btn_save.configure(text="Güncelle", fg_color="#e67e22")

                e_name.delete(0, 'end')
                e_name.insert(0, curr_name)

                e_sets.delete(0, 'end')
                e_sets.insert(0, str(curr_s))

                e_reps.delete(0, 'end')
                e_reps.insert(0, str(curr_r))

                e_cal.delete(0, 'end')
                e_cal.insert(0, str(curr_c))

            ctk.CTkButton(row, text="Kütüphaneden Sil", fg_color="#c0392b", width=120, command=lambda i=ex_id: delete_ex(i)).pack(side="right", padx=5, pady=5)
            ctk.CTkButton(row, text="✏️ Düzenle", fg_color="#d35400", width=90, command=lambda i=ex_id, n=name, set_v=s, rep_v=r, cal_v=c: edit_ex(i, n, set_v, rep_v, cal_v)).pack(side="right", padx=5, pady=5)

    # --- DETAYLI ANALİZ ---
    def show_analytics(self):
        self.clear_main_frame()
        ctk.CTkLabel(self.main_frame, text="📊 Detaylı Antrenman & Performans Analizi", font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=10, padx=10)

        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT completed, COUNT(*) FROM workout_logs GROUP BY completed")
        stats = dict(cursor.fetchall())
        completed_count = stats.get(1, 0)
        skipped_count = stats.get(-1, 0)

        cursor.execute("SELECT SUM(total_calories) FROM workout_logs WHERE completed = 1")
        total_ex_cal = cursor.fetchone()[0] or 0

        cursor.execute("SELECT SUM(walk_calories), SUM(water_l), COUNT(DISTINCT date) FROM daily_logs")
        daily_stats = cursor.fetchone()
        total_walk_cal = daily_stats[0] or 0
        total_water = daily_stats[1] or 0.0
        total_days = daily_stats[2] or 1
        conn.close()

        total_calories = total_ex_cal + total_walk_cal
        avg_calories = int(total_calories / max(total_days, 1))

        m_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        m_frame.pack(fill="x", pady=5)
        m_frame.columnconfigure((0, 1, 2), weight=1)

        self.create_metric_card(m_frame, 0, 0, "✅ Tamamlanan Antrenman", f"{completed_count} Adet", "#2ecc71")
        self.create_metric_card(m_frame, 0, 1, "❌ Atlanan Antrenman", f"{skipped_count} Adet", "#e74c3c")
        self.create_metric_card(m_frame, 0, 2, "🔥 Toplam Harcanan Kalori", f"{total_calories} kcal", "#e67e22")

        m_frame2 = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        m_frame2.pack(fill="x", pady=5)
        m_frame2.columnconfigure((0, 1), weight=1)

        self.create_metric_card(m_frame2, 0, 0, "⚡ Günlük Ortalama Kalori", f"{avg_calories} kcal/gün", "#f1c40f")
        self.create_metric_card(m_frame2, 0, 1, "💧 Toplam Tüketilen Su", f"{round(total_water, 2)} Liter", "#3498db")

    # --- EXCEL DIŞA AKTARMA ---
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Verileri Excel Olarak Kaydet")
        if not file_path: return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws_daily = wb.active
            ws_daily.title = "Daily Logs"
            ws_daily.append(["Tarih", "Yürüyüş Dk", "Adım", "Yürüyüş Kalori", "Su (L)", "Kilo (kg)", "Not"])

            conn = sqlite3.connect(db.DB_NAME)
            cursor = conn.cursor()
            cursor.execute("SELECT date, walk_minutes, walk_steps, walk_calories, water_l, weight, note FROM daily_logs")
            for row in cursor.fetchall(): ws_daily.append(list(row))

            ws_workout = wb.create_sheet(title="Workout Logs")
            ws_workout.append(["Tarih", "Egzersiz Adı", "Set", "Tekrar", "kcal/tekrar", "Toplam Kalori", "Durum"])
            cursor.execute("SELECT date, exercise_name, sets, reps, cal_per_rep, total_calories, completed FROM workout_logs")
            for row in cursor.fetchall(): ws_workout.append(list(row))

            conn.close()
            wb.save(file_path)
            messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel'e aktarma hatası:\n{str(e)}")

    # --- EXCEL İÇERİ AKTARMA ---
    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")], title="Excel Dosyası Seç")
        if not file_path: return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            conn = sqlite3.connect(db.DB_NAME)
            cursor = conn.cursor()

            if "Daily Logs" in wb.sheetnames:
                for row in wb["Daily Logs"].iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        cursor.execute('''INSERT OR REPLACE INTO daily_logs 
                                          (date, walk_minutes, walk_steps, walk_calories, water_l, weight, note)
                                          VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                       (str(row[0]), row[1], row[2], row[3], row[4], row[5], row[6]))

            if "Workout Logs" in wb.sheetnames:
                for row in wb["Workout Logs"].iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        cursor.execute('''INSERT OR REPLACE INTO workout_logs 
                                          (date, exercise_name, sets, reps, cal_per_rep, total_calories, completed)
                                          VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                       (str(row[0]), str(row[1]), row[2], row[3], row[4], row[5], row[6]))

            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı", "Excel verileri yüklendi!")
            self.show_dashboard()
        except Exception as e:
            messagebox.showerror("Hata", f"İçeri aktarma hatası:\n{str(e)}")

if __name__ == "__main__":
    app = FitnessDashboard()
    app.mainloop()